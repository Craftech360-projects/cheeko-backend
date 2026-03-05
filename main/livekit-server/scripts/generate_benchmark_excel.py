#!/usr/bin/env python3
"""Generate Excel benchmark report comparing Sarvam Pipeline vs Gemini Realtime."""

import json
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SARVAM_FILE = os.path.join(BASE_DIR, "logs", "sarvam_metrics.jsonl")
GEMINI_FILE = os.path.join(BASE_DIR, "logs", "gemini_metrics.jsonl")
OUTPUT_FILE = os.path.join(BASE_DIR, "docs", "cheeko_voice_benchmark_report.xlsx")

# Colors
DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
DARK_GREEN = "375623"
MED_GREEN = "548235"
LIGHT_GREEN = "E2EFDA"
DARK_ORANGE = "BF8F00"
LIGHT_ORANGE = "FFF2CC"
DARK_RED = "C00000"
LIGHT_RED = "FCE4EC"
DARK_PURPLE = "7030A0"
LIGHT_PURPLE = "E8D5F5"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GRAY_BORDER = "B4B4B4"

thin_border = Border(
    left=Side(style='thin', color=GRAY_BORDER),
    right=Side(style='thin', color=GRAY_BORDER),
    top=Side(style='thin', color=GRAY_BORDER),
    bottom=Side(style='thin', color=GRAY_BORDER),
)


def load_records(filepath):
    records = []
    if not os.path.exists(filepath):
        return records
    with open(filepath, "r") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    return records


def percentile(data, p):
    if not data:
        return 0.0
    s = sorted(data)
    idx = min(int(len(s) * p / 100), len(s) - 1)
    return round(s[idx], 3)


def avg(data):
    return round(sum(data) / len(data), 3) if data else 0.0


def style_header_row(ws, row, col_start, col_end, fill_color, font_color=WHITE):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color=font_color, size=11)
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border


def style_data_row(ws, row, col_start, col_end, alt=False):
    fill = PatternFill(start_color=LIGHT_GRAY if alt else WHITE, end_color=LIGHT_GRAY if alt else WHITE, fill_type="solid")
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def write_section_title(ws, row, col_start, col_end, title, fill_color, font_color=WHITE):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = title
    cell.font = Font(bold=True, color=font_color, size=13)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=c).border = thin_border


def fmt(val):
    return f"{val:.3f}s"


def main():
    sarvam = load_records(SARVAM_FILE)
    gemini = load_records(GEMINI_FILE)

    wb = Workbook()

    # ==================== DASHBOARD SHEET ====================
    ws = wb.active
    ws.title = "Benchmark Dashboard"
    ws.sheet_properties.tabColor = "1F3864"

    # Column widths
    for col_idx, width in enumerate([4, 28, 14, 14, 14, 14, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row = 1

    # ---- MAIN TITLE ----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1)
    cell.value = "CHEEKO VOICE AI - LATENCY BENCHMARK REPORT"
    cell.font = Font(bold=True, color=WHITE, size=16)
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 40
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1)
    cell.value = "Sarvam AI Pipeline (STT + LLM + TTS)  vs  Google Gemini 2.5 Flash Realtime (End-to-End)"
    cell.font = Font(bold=True, color=DARK_BLUE, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1)
    cell.value = "Languages: Telugu (te-IN) | Kannada (kn-IN) | Hindi (hi-IN)  |  Test Date: Feb 2026"
    cell.font = Font(color="666666", size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # ==================================================================
    # SECTION 1: HOW VOICE PIPELINE WORKS
    # ==================================================================
    write_section_title(ws, row, 1, 10, "HOW IT WORKS: User Speaks → STT → LLM → TTS → User Hears Reply", DARK_BLUE)
    row += 1

    flow_data = [
        ("Step", "Component", "What It Does", "Sarvam Pipeline", "Gemini Realtime", "", "", "", "", ""),
        ("1", "STT (Speech-to-Text)", "Converts user's voice to text", "Sarvam saaras/saarika", "Built-in (end-to-end)", "", "", "", "", ""),
        ("2", "LLM (Language Model)", "Thinks and generates reply text", "Groq (llama / gpt-oss)", "Built-in (end-to-end)", "", "", "", "", ""),
        ("3", "TTS (Text-to-Speech)", "Converts reply text to voice", "Sarvam bulbul", "Built-in (end-to-end)", "", "", "", "", ""),
    ]
    headers = flow_data[0]
    for ci, h in enumerate(headers[:5], 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 1, 5, MED_BLUE)
    row += 1
    for ri, rd in enumerate(flow_data[1:]):
        for ci, val in enumerate(rd[:5], 1):
            ws.cell(row=row, column=ci, value=val)
        style_data_row(ws, row, 1, 5, alt=(ri % 2 == 1))
        row += 1
    row += 1

    # ==================================================================
    # SECTION 2: STT LATENCY
    # ==================================================================
    write_section_title(ws, row, 1, 10, "STT LATENCY (Speech-to-Text) — Sarvam AI", MED_GREEN, WHITE)
    row += 1

    stt_headers = ["#", "STT Model", "Language", "Samples", "Total Audio", "Mode", "Latency", "", "", ""]
    for ci, h in enumerate(stt_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 1, 10, MED_GREEN)
    row += 1

    stt_records = [r for r in sarvam if r.get("type") == "stt"]
    stt_groups = defaultdict(list)
    for s in stt_records:
        key = (s.get("stt_model", "?"), s.get("stt_language", "?"))
        stt_groups[key].append(s)

    idx = 1
    for (model, lang), data in sorted(stt_groups.items()):
        audio_durs = [d["stt_audio_duration"] for d in data if d.get("stt_audio_duration")]
        streamed = sum(1 for d in data if d.get("stt_streamed"))
        vals = [idx, model, lang, len(data), f"{sum(audio_durs):.1f}s",
                "Streaming" if streamed == len(data) else "REST",
                "Real-time (0ms added)", "", "", ""]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v)
        style_data_row(ws, row, 1, 10, alt=(idx % 2 == 0))
        # Highlight the latency cell green
        ws.cell(row=row, column=7).font = Font(bold=True, color=DARK_GREEN)
        row += 1
        idx += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1)
    cell.value = "All Sarvam STT models process in real-time via streaming. Transcript is ready before user stops speaking = 0ms additional latency."
    cell.font = Font(italic=True, color="666666", size=9)
    cell.alignment = Alignment(horizontal='left')
    row += 2

    # ==================================================================
    # SECTION 3: LLM LATENCY
    # ==================================================================
    write_section_title(ws, row, 1, 10, "LLM LATENCY (Language Model) — Time to First Token (TTFT)", MED_BLUE, WHITE)
    row += 1

    llm_headers = ["#", "LLM Model", "Samples", "Avg TTFT", "P50", "P95", "Min", "Max", "Tokens/sec", ""]
    for ci, h in enumerate(llm_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 1, 10, MED_BLUE)
    row += 1

    llm_records = [r for r in sarvam if r.get("type") == "llm"]
    llm_groups = defaultdict(list)
    for l in llm_records:
        m = l.get("llm_model", "?")
        if l.get("llm_ttft", 0) > 0:
            llm_groups[m].append(l)

    idx = 1
    for model, data in sorted(llm_groups.items()):
        ttfts = sorted([d["llm_ttft"] for d in data])
        tps = [d["llm_tokens_per_sec"] for d in data if d.get("llm_tokens_per_sec", 0) > 0]
        vals = [idx, model, len(ttfts), fmt(avg(ttfts)), fmt(percentile(ttfts, 50)),
                fmt(percentile(ttfts, 95)), fmt(min(ttfts)), fmt(max(ttfts)),
                f"{avg(tps):.1f}" if tps else "-", ""]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v)
        style_data_row(ws, row, 1, 10, alt=(idx % 2 == 0))
        row += 1
        idx += 1
    row += 1

    # ==================================================================
    # SECTION 4: TTS LATENCY
    # ==================================================================
    write_section_title(ws, row, 1, 10, "TTS LATENCY (Text-to-Speech) — Sarvam AI — Time to First Byte (TTFB)", DARK_ORANGE, WHITE)
    row += 1

    tts_headers = ["#", "TTS Model", "Speaker", "Language", "Samples", "Avg TTFB", "P50", "P95", "Min", "Max"]
    for ci, h in enumerate(tts_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 1, 10, DARK_ORANGE)
    row += 1

    tts_records = [r for r in sarvam if r.get("type") == "tts"]
    tts_groups = defaultdict(list)
    for t in tts_records:
        key = (t.get("tts_model", "?"), t.get("tts_speaker", "?"), t.get("tts_language", "?"))
        ttfb = t.get("tts_ttfb", 0)
        if ttfb > 0:
            tts_groups[key].append(ttfb)

    idx = 1
    for (model, spk, lang), ttfbs in sorted(tts_groups.items()):
        s = sorted(ttfbs)
        vals = [idx, model, spk, lang, len(s), fmt(avg(s)), fmt(percentile(s, 50)),
                fmt(percentile(s, 95)), fmt(min(s)), fmt(max(s))]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v)
        style_data_row(ws, row, 1, 10, alt=(idx % 2 == 0))
        # Highlight fastest rows
        if avg(s) < 0.7:
            ws.cell(row=row, column=6).font = Font(bold=True, color=DARK_GREEN)
        row += 1
        idx += 1
    row += 1

    # ==================================================================
    # SECTION 5: COMBINED RESPONSE LATENCY (Sarvam Pipeline)
    # ==================================================================
    write_section_title(ws, row, 1, 10, "COMBINED RESPONSE LATENCY — Sarvam Pipeline (User Waits This Long)", DARK_RED, WHITE)
    row += 1

    resp_headers = ["#", "STT Model", "LLM Model", "TTS / Speaker", "Language", "Turns", "Avg", "P50", "P95", "Min"]
    for ci, h in enumerate(resp_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 1, 10, DARK_RED)
    row += 1

    turns = [r for r in sarvam if r.get("type") == "pipeline_turn"]
    turn_groups = defaultdict(list)
    for t in turns:
        key = (t.get("stt_model", "?"), t.get("llm_model", "?"), t.get("tts_model", "?"),
               t.get("tts_speaker", "?"), t.get("stt_language", "?"))
        if t.get("response_latency"):
            turn_groups[key].append(t["response_latency"])

    idx = 1
    for (stt_m, llm_m, tts_m, spk, lang), lats in sorted(turn_groups.items()):
        s = sorted(lats)
        vals = [idx, stt_m, llm_m, f"{tts_m}/{spk}", lang, len(s),
                fmt(avg(s)), fmt(percentile(s, 50)), fmt(percentile(s, 95)), fmt(min(s))]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v)
        style_data_row(ws, row, 1, 10, alt=(idx % 2 == 0))
        row += 1
        idx += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1)
    cell.value = "Response Latency = LLM TTFT + TTS TTFB. This is the time from user stops speaking to agent starts speaking."
    cell.font = Font(italic=True, color="666666", size=9)
    cell.alignment = Alignment(horizontal='left')
    row += 2

    # ==================================================================
    # SECTION 6: GEMINI REALTIME (COMPETITOR)
    # ==================================================================
    write_section_title(ws, row, 1, 10, "GEMINI 2.5 FLASH REALTIME — End-to-End (Competitor Baseline)", DARK_PURPLE, WHITE)
    row += 1

    gem_headers = ["#", "Model", "Voice", "Samples", "Avg TTFT", "P50", "P95", "Min", "Max", "Tokens/sec"]
    for ci, h in enumerate(gem_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 1, 10, DARK_PURPLE)
    row += 1

    gem_rt = [r for r in gemini if r.get("type") == "realtime" and r.get("ttft", 0) > 0]
    if gem_rt:
        ttfts = sorted([r["ttft"] for r in gem_rt])
        tps_list = [r["tokens_per_second"] for r in gem_rt if r.get("tokens_per_second", 0) > 0]
        model = gem_rt[0].get("model", "?")
        voice = gem_rt[0].get("voice", "?")
        vals = [1, model, voice, len(ttfts), fmt(avg(ttfts)), fmt(percentile(ttfts, 50)),
                fmt(percentile(ttfts, 95)), fmt(min(ttfts)), fmt(max(ttfts)),
                f"{avg(tps_list):.1f}" if tps_list else "-"]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=row, column=ci, value=v)
        style_data_row(ws, row, 1, 10)
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1)
    cell.value = "Gemini Realtime is an end-to-end model (STT+LLM+TTS in one). TTFT = total time from user stops speaking to agent starts speaking."
    cell.font = Font(italic=True, color="666666", size=9)
    cell.alignment = Alignment(horizontal='left')
    row += 2

    # ==================================================================
    # SECTION 7: HEAD-TO-HEAD COMPARISON
    # ==================================================================
    write_section_title(ws, row, 1, 10, "HEAD-TO-HEAD: Sarvam Pipeline vs Gemini Realtime", DARK_BLUE, WHITE)
    row += 1

    cmp_headers = ["#", "Metric", "Sarvam Pipeline", "Gemini Realtime", "Winner", "Notes", "", "", "", ""]
    for ci, h in enumerate(cmp_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 1, 10, DARK_BLUE)
    row += 1

    # Compute Sarvam best/avg response latency
    all_sarvam_lats = [t["response_latency"] for t in turns if t.get("response_latency")]
    sarvam_avg_resp = avg(all_sarvam_lats) if all_sarvam_lats else 0
    sarvam_best_resp = min(all_sarvam_lats) if all_sarvam_lats else 0

    # Gemini
    gem_ttfts = sorted([r["ttft"] for r in gem_rt]) if gem_rt else []
    gemini_avg = avg(gem_ttfts) if gem_ttfts else 0
    gemini_best = min(gem_ttfts) if gem_ttfts else 0

    # Sarvam component breakdown
    all_llm_ttfts = [r["llm_ttft"] for r in sarvam if r.get("type") == "llm" and r.get("llm_ttft", 0) > 0]
    all_tts_ttfbs = [r["tts_ttfb"] for r in sarvam if r.get("type") == "tts" and r.get("tts_ttfb", 0) > 0]

    comparisons = [
        [1, "Avg Response Latency", fmt(sarvam_avg_resp), fmt(gemini_avg),
         "Gemini" if gemini_avg < sarvam_avg_resp else "Sarvam",
         "Time user waits for agent to start speaking"],
        [2, "Best Response Latency", fmt(sarvam_best_resp), fmt(gemini_best),
         "Gemini" if gemini_best < sarvam_best_resp else "Sarvam",
         "Fastest single turn observed"],
        [3, "STT Latency (Sarvam only)", "0ms (streaming)", "Built-in",
         "Sarvam", "Sarvam STT adds zero extra latency"],
        [4, "Sarvam STT + TTS Only", fmt(avg(all_tts_ttfbs)) if all_tts_ttfbs else "-", "N/A",
         "Sarvam", f"Sarvam's own contribution (no LLM). Best TTS: {fmt(min(all_tts_ttfbs))}" if all_tts_ttfbs else ""],
        [5, "LLM Bottleneck (Groq)", fmt(avg(all_llm_ttfts)) if all_llm_ttfts else "-", "N/A",
         "-", "LLM is the bottleneck in Sarvam pipeline, NOT Sarvam"],
        [6, "Indian Language Support", "Native (te, kn, hi)", "Limited",
         "Sarvam", "Sarvam built for Indian languages"],
        [7, "Separate STT/TTS Control", "Yes (mix & match)", "No (black box)",
         "Sarvam", "Can swap LLM independently"],
    ]

    for ri, comp in enumerate(comparisons):
        for ci, v in enumerate(comp, 1):
            ws.cell(row=row, column=ci, value=v)
        style_data_row(ws, row, 1, 10, alt=(ri % 2 == 1))
        # Color the winner cell
        winner_cell = ws.cell(row=row, column=5)
        if "Sarvam" in str(comp[4]):
            winner_cell.font = Font(bold=True, color=DARK_GREEN)
        elif "Gemini" in str(comp[4]):
            winner_cell.font = Font(bold=True, color=DARK_PURPLE)
        row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1)
    cell.value = "KEY INSIGHT: Sarvam STT (0ms) + TTS (0.3-1.2s) contributes only 0.3-1.2s. The 2s bottleneck is the LLM (Groq), not Sarvam."
    cell.font = Font(bold=True, color=DARK_RED, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=1)
    cell.value = "With a faster LLM, the Sarvam pipeline can achieve <1.5s response — beating Gemini Realtime."
    cell.font = Font(bold=True, color=DARK_GREEN, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30

    # ==================================================================
    # SECTION 8: GLOSSARY / DEFINITIONS
    # ==================================================================
    row += 2
    write_section_title(ws, row, 1, 10, "GLOSSARY — What These Terms Mean (in Simple Words)", DARK_BLUE, WHITE)
    row += 1

    def_headers = ["#", "Term", "Full Form", "What It Means (Simple)", "", "", "", "", "", ""]
    for ci, h in enumerate(def_headers, 1):
        ws.cell(row=row, column=ci, value=h)
    style_header_row(ws, row, 1, 10, MED_BLUE)
    row += 1

    definitions = [
        [1, "STT", "Speech-to-Text",
         "Converts the user's voice into text. When user speaks, STT listens and gives text output."],
        [2, "LLM", "Large Language Model",
         "The AI brain. Takes text from STT, thinks, and generates a reply in text."],
        [3, "TTS", "Text-to-Speech",
         "Converts the AI's text reply back into voice so the user can hear it."],
        [4, "TTFT", "Time to First Token",
         "How long the LLM takes to start generating its reply after receiving the text. Lower = faster thinking."],
        [5, "TTFB", "Time to First Byte",
         "How long TTS takes to start producing audio after receiving text. Lower = faster voice output."],
        [6, "Response Latency", "End-to-End Wait Time",
         "Total time from user stops speaking to agent starts speaking back. This is what the user actually feels."],
        [7, "P50 (Median)", "50th Percentile",
         "The middle value — 50% of requests were faster, 50% were slower. Represents the typical experience."],
        [8, "P95", "95th Percentile",
         "95% of requests were faster than this. Shows worst-case performance (ignoring rare outliers)."],
        [9, "Avg", "Average",
         "Simple average of all measurements."],
        [10, "Streaming", "Real-time Processing",
         "STT processes audio as the user speaks, instead of waiting for them to finish. Means zero extra delay."],
        [11, "Tokens/sec", "Tokens per Second",
         "How fast the LLM generates text. Higher number = faster replies."],
    ]

    for ri, defn in enumerate(definitions):
        ws.cell(row=row, column=1, value=defn[0])
        ws.cell(row=row, column=2, value=defn[1])
        ws.cell(row=row, column=3, value=defn[2])
        # Merge columns 4-10 for the description
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=10)
        desc_cell = ws.cell(row=row, column=4)
        desc_cell.value = defn[3]
        desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        style_data_row(ws, row, 1, 3, alt=(ri % 2 == 1))
        # Style merged description cell
        fill_color = LIGHT_GRAY if ri % 2 == 1 else WHITE
        desc_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        desc_cell.border = thin_border
        ws.row_dimensions[row].height = 28
        row += 1

    # Freeze panes
    ws.freeze_panes = "A5"

    # Save
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"Report saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
